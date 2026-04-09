"""Generate a blank Excel template for CityGML Energy input.

Creates an .xlsx workbook with:
* A ``_settings`` sheet for city-model metadata and geometry sources.
* One sheet per supported feature type, with canonical attribute names in
  column A (fill in instances column-by-column starting from B).

Usage::

    python tools/generate_excel_template.py [output_path]

Default output: ``inputs/citygml_energy_template.xlsx``
"""

from __future__ import annotations

import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from citygml_energy.input_catalog import (
    FEATURE_INPUT_FIELDS,
    list_supported_feature_types,
)

DEFAULT_OUTPUT = REPO_ROOT / "inputs" / "citygml_energy_template.xlsx"

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_SETTINGS_KEY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def generate_template(output_path: Path) -> None:
    wb = Workbook()

    # ── _settings sheet ────────────────────────────────────────────────
    ws_settings = wb.active
    ws_settings.title = "_settings"  # type: ignore[union-attr]
    ws_settings.column_dimensions["A"].width = 40  # type: ignore[union-attr]
    ws_settings.column_dimensions["B"].width = 60  # type: ignore[union-attr]

    settings_rows = [
        ("name", ""),
        ("description", ""),
        ("", ""),
        ("geometry_source.type", "step-renodat-lod3"),
        ("geometry_source.path", ""),
        ("geometry_source.target_building_id", ""),
        ("geometry_source.target_pv_id", ""),
    ]
    for row_idx, (key, value) in enumerate(settings_rows, start=1):
        cell_a = ws_settings.cell(row=row_idx, column=1, value=key)  # type: ignore[union-attr]
        ws_settings.cell(row=row_idx, column=2, value=value)  # type: ignore[union-attr]
        if key:
            cell_a.font = _HEADER_FONT
            cell_a.fill = _SETTINGS_KEY_FILL

    # ── Feature type sheets (transposed: col A = attr names, cols B+ = instances) ──
    for feature_type in list_supported_feature_types():
        fields = FEATURE_INPUT_FIELDS[feature_type]
        # Collect canonical names, ensuring gml_id is first
        headers: list[str] = []
        seen: set[str] = set()
        for priority_key in ("gml_id", "gml_parent_id", "gml_description", "gml_name"):
            for field in fields:
                if field.canonical == priority_key and priority_key not in seen:
                    headers.append(priority_key)
                    seen.add(priority_key)

        for field in fields:
            if field.canonical not in seen:
                headers.append(field.canonical)
                seen.add(field.canonical)

        # Excel limits sheet names to 31 chars
        sheet_title = feature_type[:31]
        ws = wb.create_sheet(title=sheet_title)

        # Write attribute names down column A
        for row_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=1, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL

        # Size column A to fit the longest attribute name
        max_len = max((len(h) for h in headers), default=14)
        ws.column_dimensions["A"].width = max_len + 4

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def main() -> None:
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT
    generate_template(output)
    print(f"Template written to {output}")


if __name__ == "__main__":
    main()
