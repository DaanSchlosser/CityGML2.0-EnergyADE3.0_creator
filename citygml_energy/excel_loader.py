"""Load CityGML feature collections from Excel workbooks.

Excel layout (transposed — attribute names down column A):

* **``_settings`` sheet** — two-column key/value table for city-model metadata
  (``name``, ``description``) and geometry sources.
* **One sheet per feature type** — sheet name = feature type
  (e.g. ``bldg_Building``, ``nrg3_PhotovoltaicCollector``).
  Column A contains attribute names (canonical or FME alias).
  Each subsequent column (B, C, D, ...) is one feature instance.

Example ``bldg_Building`` sheet::

    |  A (attribute)       |  B (instance 1)     |  C (instance 2) |
    |----------------------|---------------------|------------------|
    |  gml_id              |  id_building_1      |  id_building_2   |
    |  gml_name            |  Han solo's house   |  ...             |
    |  bldg_class          |  1000               |  ...             |

Geometry sources are encoded as rows in the ``_settings`` sheet with
``geometry_source.*`` keys::

    geometry_source.type                step-renodat-lod3
    geometry_source.path                Owner-Occupier1_LOD3.0_STEP.stp
    geometry_source.target_building_id  id_building_1
    geometry_source.target_pv_id        pv_panel_1
    ---                                 (separator — starts a new geometry source)

Multiple geometry sources are separated by a row whose key is ``---``.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .core import CityModel
from .input_loader import InputFileError, build_city_model_from_feature_collection

PathLike = str | Path

_SETTINGS_SHEET = "_settings"
_GEOMETRY_SOURCE_PREFIX = "geometry_source."
_GEOMETRY_SOURCE_SEPARATOR = "---"

_CITY_MODEL_KEYS = {"name", "description"}


def load_excel_feature_collection(path: PathLike) -> dict[str, Any]:
    """Load and convert an Excel workbook into the canonical feature-collection dict.

    The returned dict has the same shape as the JSON format and can be passed
    directly to ``build_city_model_from_feature_collection``.
    """
    input_path = Path(path)
    if not input_path.is_file():
        raise InputFileError(f"Input file not found: {input_path}")

    wb = load_workbook(str(input_path), read_only=True, data_only=True)
    try:
        return _workbook_to_feature_collection(wb, source=str(input_path))
    finally:
        wb.close()


def load_city_model_from_excel(path: PathLike) -> CityModel:
    """Load, validate, and build a CityModel from an Excel workbook."""
    input_path = Path(path)
    data = load_excel_feature_collection(input_path)
    return build_city_model_from_feature_collection(data, base_path=input_path.parent)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _workbook_to_feature_collection(wb: Any, *, source: str) -> dict[str, Any]:
    """Convert an openpyxl Workbook into the canonical dict format."""
    city_model: dict[str, str] = {}
    geometry_sources: list[dict[str, str]] = []
    features: list[dict[str, Any]] = []

    # ── Parse _settings sheet ──────────────────────────────────────────
    if _SETTINGS_SHEET in wb.sheetnames:
        settings_ws = wb[_SETTINGS_SHEET]
        current_geo: dict[str, str] = {}
        for row in settings_ws.iter_rows(min_row=1, values_only=True):
            key = _cell_str(row[0]) if len(row) > 0 else ""
            value = _cell_str(row[1]) if len(row) > 1 else ""
            if not key:
                continue

            if key == _GEOMETRY_SOURCE_SEPARATOR:
                if current_geo:
                    geometry_sources.append(current_geo)
                    current_geo = {}
                continue

            if key.startswith(_GEOMETRY_SOURCE_PREFIX):
                geo_key = key[len(_GEOMETRY_SOURCE_PREFIX) :]
                if value:
                    current_geo[geo_key] = value
            elif key in _CITY_MODEL_KEYS:
                if value:
                    city_model[key] = value
            else:
                raise InputFileError(f"{source}: _settings sheet contains unknown key {key!r}")

        # Flush last geometry source if any
        if current_geo:
            geometry_sources.append(current_geo)

    # ── Parse feature sheets (transposed: col A = headers, cols B+ = instances) ──
    for sheet_name in wb.sheetnames:
        if sheet_name == _SETTINGS_SHEET:
            continue

        feature_type = sheet_name
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Column A of each row is the attribute name
        attr_names = [_cell_str(row[0]) if row else "" for row in rows]

        # Determine how many instance columns exist (B, C, D, ...)
        max_cols = max((len(row) for row in rows), default=1)
        if max_cols < 2:
            continue  # no data columns

        for col_idx in range(1, max_cols):
            attrs = _column_to_attrs(attr_names, rows, col_idx)
            if not attrs:
                continue
            features.append(
                {
                    "feature_type": feature_type,
                    "attributes": attrs,
                }
            )

    result: dict[str, Any] = {
        "schema_version": 1,
        "city_model": city_model,
        "features": features,
    }
    if geometry_sources:
        result["geometry_sources"] = geometry_sources

    return result


def _column_to_attrs(
    attr_names: list[str],
    rows: list[tuple[Any, ...]],
    col_idx: int,
) -> dict[str, Any]:
    """Extract one feature's attributes from a single column."""
    attrs: dict[str, Any] = {}
    for row_idx, name in enumerate(attr_names):
        if not name:
            continue
        row = rows[row_idx]
        value = row[col_idx] if col_idx < len(row) else None
        if value is None:
            continue
        if isinstance(value, float) and value == int(value):
            value = int(value)
        elif isinstance(value, str) and value.strip() == "":
            continue
        attrs[name] = value
    return attrs


def _cell_str(value: Any) -> str:
    """Convert a cell value to a stripped string, or empty string if None."""
    if value is None:
        return ""
    return str(value).strip()
